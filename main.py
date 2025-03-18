import os
import PyPDF2
import re
import pandas as pd
from openpyxl import load_workbook
import shutil

DIST = 'NATURGY'

class ExtratorFaturas:
    def __init__(self):
        self.regexes = {
            'cnpj': [r'\-\s(\d+\.\d+\.\d+\/\d+\-\d+)\s?'], #05/09
            'valor_total': [r'\-\d+\s(\d+\.?\,?\d+\,\d{2}?)\s\d+', r'\d+\s(\d+\.?\,?\d+\.?\,?\d+\,?\.?\d{2}?)\s\d+'], #10/09
            'volume_total': [r'TRIBUTOS(\s?\d+\.?\,?\d+\.?\,?\d+?\.?\,?\d+?)\s?'], #10/09
            'data_emissao': [r'\d+?\s?(\d{2}\/\d{2}\/\d{4})\s?'], #10/09
            'data_inicio': [r'\s?(\d{2}\/\d{2}\/\d{4})\s[Aa?]', r'\d{2}\/\d{2}\/\d{4}\s?(\d{2}\/\d{2}\/\d{4})\s?\d{2}\/\d{2}\/\d{4}'], #11/09
            'data_fim': [r'\s?\d{2}\/\d{2}\/\d{4}\s[Aa?]\s(\d{2}\/\d{2}\/\d{4})', r'\d{2}\/\d{2}\/\d{4}\s?\d{2}\/\d{2}\/\d{4}\s?(\d{2}\/\d{2}\/\d{4})'], #11/09            
            'numero_documento': [r'[F](\d{2}\s\d+)'],
            'valor_icms': [r'\d+\.\d+?\,?\d+\,\d{2}\s(\d+\.\d+?\,?\d+\,\d{2})'], #11/09     
            'correcao_pcs': [r'(\d+\,?\.?\d+)\s\d+\.\d+?\,?\d+\,\d{2}\sFORNECIMENTO', r'(\d+\,?\.?\d+)\s\.?\,?\d+\.?\,?\d+\.?\,?\d+\.?\,?\d+\sFORNECIMENTO'], #11/09
            'codigo_cliente': [r'\d+?\-\d{3}(\d+\-?\d+?)\s'] #12/09
        }

    def extrair_informacoes(self, texto):
        informacoes = {}
        for chave, regex_list in self.regexes.items():
            for regex in regex_list:
                match = re.search(regex, texto)
                if match:
                    informacoes[chave] = match.group(1) if match.groups() else match.group(0)
                    break  # Para de procurar assim que encontrar uma correspondência
            if chave not in informacoes:
                print(f"❌ Valor não encontrado para: {chave}")
                return None  # Retorna None se algum valor não for encontrado
        return informacoes

def extrair_texto(caminho_do_pdf):
    texto_completo = ''
    try:
        with open(caminho_do_pdf, 'rb') as arquivo:
            leitor_pdf = PyPDF2.PdfReader(arquivo)
            
            # Itera por todas as páginas do PDF
            for pagina in leitor_pdf.pages:
                texto_pagina = pagina.extract_text()
                
                if texto_pagina:
                    texto_pagina = texto_pagina.replace('\n', ' ')
                    texto_pagina = re.sub(r'\s{1,2,}', ' ', texto_pagina).strip()
                    texto_completo += texto_pagina + ' '
                else:
                    print(f"⚠️ Não foi possível extrair texto de uma página do PDF: {caminho_do_pdf}")
            
            # Remove espaços extras no início e no final
            texto_completo = texto_completo.strip()
            
            if not texto_completo:
                print(f"❌ Nenhum texto extraído do PDF {caminho_do_pdf}")
            else:
                print(f"✅ Texto extraído com sucesso...")  # Mostra os primeiros 500 caracteres
    
    except Exception as e:
        print(f"❌ Erro ao abrir ou ler o PDF {caminho_do_pdf}: {e}")
        texto_completo = ''
    
    return texto_completo

def registro_existe(df, cnpj, data_inicio, data_fim, valor_total):
    return not df[(df['CNPJ'] == cnpj) & (df['DATA INICIO'] == data_inicio) & (df['DATA FIM'] == data_fim) & (df['VALOR TOTAL'] == valor_total)].empty

def todos_campos_preenchidos(informacoes):
    campos_obrigatorios = ['cnpj', 'valor_total', 'volume_total', 'data_emissao', 'data_inicio', 'data_fim', 'numero_documento', 'valor_icms', 'correcao_pcs', 'codigo_cliente']
    for campo in campos_obrigatorios:
        if campo not in informacoes or not informacoes[campo]:
            print(f"❌ Campo obrigatório '{campo}' está faltando ou vazio.")
            return False
    return True

def adicionar_na_planilha(informacoes, caminho_planilha, nome_arquivo):
    if not todos_campos_preenchidos(informacoes):
        print("❌ Não foi possível adicionar à planilha devido a campos faltantes ou vazios.")
        return False

    try:
        df = pd.read_excel(caminho_planilha)
    except FileNotFoundError:
        print(f"ℹ️ O arquivo '{caminho_planilha}' não foi encontrado. Criando um novo.")
        df = pd.DataFrame(columns=['CNPJ', 'VALOR TOTAL', 'VOLUME TOTAL', 'DATA EMISSAO', 'DATA INICIO', 'DATA FIM', 'NUMERO FATURA', 'VALOR ICMS', 'CORRECAO PCS', 'CODIGO CLIENTE', 'DISTRIBUIDORA', 'NOME DO ARQUIVO'])
    
    cnpj = informacoes['cnpj']
    data_inicio = informacoes['data_inicio']
    data_fim = informacoes['data_fim']
    valor_total = pd.to_numeric(informacoes['valor_total'].replace('.', '').replace(',', '.'))
    volume_total = pd.to_numeric(informacoes['volume_total'].replace('.', '').replace(',', '.'))
    valor_icms = pd.to_numeric(informacoes['valor_icms'].replace('.', '').replace(',', '.'))

    if registro_existe(df, cnpj, data_inicio, data_fim, valor_total):
        print(f"❌ Registro duplicado encontrado. Não será inserido.")
        return False 

    valor_total = pd.to_numeric(str(informacoes.get('valor_total', '')).replace('.', '').replace(',', '.'), errors='coerce')
    volume_total = pd.to_numeric(str(informacoes.get('volume_total', '')).replace('.', '').replace(',', '.'), errors='coerce')
    valor_icms = pd.to_numeric(str(informacoes.get('valor_icms', '')).replace('.', '').replace(',', '.'), errors='coerce')
    correcao_pcs = pd.to_numeric(informacoes.get('correcao_pcs', ''), errors='coerce')    
    
    nova_linha = pd.DataFrame([{
        'CNPJ': cnpj,
        'VALOR TOTAL': valor_total,
        'VOLUME TOTAL': volume_total,
        'DATA EMISSAO': informacoes['data_emissao'],
        'DATA INICIO': data_inicio,
        'DATA FIM': data_fim,
        'NUMERO FATURA': informacoes['numero_documento'],
        'VALOR ICMS': valor_icms,
        'CORRECAO PCS': informacoes['correcao_pcs'],
        'CODIGO CLIENTE': informacoes['codigo_cliente'],
        'DISTRIBUIDORA': DIST,
        'NOME DO ARQUIVO': nome_arquivo
    }])
    df = pd.concat([df, nova_linha], ignore_index=True)
    df.to_excel(caminho_planilha, index=False)
    print("✅ Dados adicionados com sucesso à planilha.")
    return True

def mover_arquivo(origem, destino):
    shutil.move(origem, destino)
    print(f"Arquivo movido para {destino}")

def verificar_linha_preenchida(caminho_planilha, informacoes):
    try:
        workbook = load_workbook(caminho_planilha)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Ignora o cabeçalho
            if (
                row[0] == informacoes.get('cnpj') and
                row[1] == informacoes.get('valor_total') and
                row[2] == informacoes.get('volume_total') and
                row[3] == informacoes.get('data_emissao') and
                row[4] == informacoes.get('data_inicio') and
                row[5] == informacoes.get('data_fim') and
                row[6] == informacoes.get('numero_documento') and
                row[7] == informacoes.get('valor_icms') and
                row[8] == informacoes.get('correcao_pcs') and
                row[9] == informacoes.get('codigo_cliente')
            ):
                if all(cell is not None and cell != '' for cell in row):
                    return True
                else:
                    return False
        return False  # Retorna False se a linha correspondente não for encontrada
    except Exception as e:
        print(f"❌ Erro ao verificar a planilha: {e}")
        return False

def main(file_path, pdf_file, caminho_planilha):
    texto_pypdf = extrair_texto(pdf_file)
    if not texto_pypdf:
        print(f"❌ Erro ao extrair texto do PDF: {pdf_file}")
        return

    extrator = ExtratorFaturas()
    informacoes = extrator.extrair_informacoes(texto_pypdf)
    if not informacoes:
        print(f"❌ Nenhuma informação extraída do PDF: {pdf_file}")
        return

    nome_arquivo = os.path.basename(pdf_file)
    inserido = adicionar_na_planilha(informacoes, caminho_planilha, nome_arquivo)
    print(informacoes)

    if inserido:
        destino = os.path.join(diretorio_destino, nome_arquivo)
        mover_arquivo(pdf_file, destino)
    else:
        print('❌ Arquivo não foi inserido na planilha devido a dados faltantes ou duplicados. Não será movido.')

# Exemplo de uso
file_path = r'G:\QUALIDADE\Códigos\Leitura de Faturas Gás\Códigos\Naturgy RJ\Faturas'
diretorio_destino = r'G:\QUALIDADE\Códigos\Leitura de Faturas Gás\Códigos\Naturgy RJ\Lidos'
caminho_planilha = r'G:\QUALIDADE\Códigos\Leitura de Faturas Gás\Códigos\00 Faturas Lidas\NATURGY.xlsx'

for arquivo in os.listdir(file_path):
    if arquivo.endswith('.pdf') or arquivo.endswith('.PDF'):
        arquivo_full = os.path.join(file_path, arquivo)
        arquivo = os.path.basename(arquivo)

        main(arquivo, arquivo_full, caminho_planilha)